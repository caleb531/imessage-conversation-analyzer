#!/usr/bin/env python3
# type: ignore

import os
import os.path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

import ica
from ica.analyzers.message_totals import get_results as message_totals


def main() -> None:
    cli_args = ica.get_cli_args()

    sheet_name = "Message Totals"
    df = ica.normalize_df_for_output(message_totals(cli_args))
    is_default_index = not df.index.name

    excel_file_path = os.path.expanduser("~/Desktop/output.xlsx")

    # Create a new Excel workbook and set up a sheet
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write DataFrame headers with styling
    for col_num, column_name in enumerate(
        df.columns, 1
    ):  # Columns start from 1 in openpyxl
        cell = ws.cell(row=1, column=col_num, value=column_name)
        # Apply header styles
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Write DataFrame rows
    for row_num, row in enumerate(
        df.itertuples(index=not is_default_index), 2
    ):  # Rows start from 2 (after headers)
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    wb.save(excel_file_path)


if __name__ == "__main__":
    main()
